import argparse
import os

import cv2
import mmcv
import numpy as np
import pycocotools.mask as maskUtils
import torch

import openpyxl
from PIL import Image, ImageFont, ImageDraw
from mmcv import color_val, imwrite
from mmdet.core.evaluation.bbox_overlaps import bbox_overlaps

from mmdet.ops import nms

from mmdet.apis import init_detector, inference_detector, show_result
from mmdet.datasets.shell import ShellDataset
from mmdet.datasets.sku import SkuDataset
from mmdet.datasets.rosegold import RoseGoldDataset
from mmdet.datasets.UltraAB import UltraABDataset
import xml.etree.ElementTree as ET

from util.py_nms import py_cpu_nms

TABLE_HEAD = ["名称", "样本个数", "tp", "fp", "fn", "precision", "recall"]

test_img_path = "/home/lichengzhi/mmdetection/data/VOCdevkit/shell/2020.04.17/JPEGImages"
test_xml_path = "/home/lichengzhi/mmdetection/data/VOCdevkit/shell/2020.04.17/Annotations"
test_path = "/home/lichengzhi/mmdetection/data/VOCdevkit/shell/2020.04.17/ImageSets/Main/test.txt"


def parse_args():
    parser = argparse.ArgumentParser(description='MMDet test detector')
    parser.add_argument('configrg', help='test config file path')
    parser.add_argument('checkpointrg', help='checkpoint file')
    parser.add_argument('configab', help='test config file path')
    parser.add_argument('checkpointab', help='checkpoint file')
    parser.add_argument(
        '--eval',
        type=str,
        nargs='+',
        choices=['proposal', 'proposal_fast', 'bbox', 'segm', 'keypoints'],
        help='eval types')
    parser.add_argument('--show', action='store_true', help='show results')
    parser.add_argument('--tmpdir', help='tmp dir for writing some results')
    parser.add_argument(
        '--launcher',
        choices=['none', 'pytorch', 'slurm', 'mpi'],
        default='none',
        help='job launcher')
    parser.add_argument('--local_rank', type=int, default=0)
    parser.add_argument('--image', help='detect image file')
    args = parser.parse_args()
    return args


def read_xml(xml_path):
    tree = ET.parse(xml_path)
    root = tree.getroot()
    objs = root.findall('object')
    coords = list()
    for ix, obj in enumerate(objs):
        name = obj.find('name').text
        box = obj.find('bndbox')
        x_min = float(box[0].text)
        y_min = float(box[1].text)
        x_max = float(box[2].text)
        y_max = float(box[3].text)
        coords.append([x_min, y_min, x_max, y_max, name])
    return coords


def get_result(resultrg, resultab, rg_class_names, ab_class_names, score_thr=0.5):
    rg_bbox_result = resultrg
    ab_bbox_result = resultab

    bboxes_rg = np.vstack(rg_bbox_result)
    bboxes_ab = np.vstack(ab_bbox_result)

    labels_rg = [
        np.full(bbox.shape[0], rg_class_names[i])
        for i, bbox in enumerate(rg_bbox_result)
    ]
    labels_rg = np.concatenate(labels_rg)

    labels_ab = [
        np.full(bbox.shape[0], ab_class_names[i])
        for i, bbox in enumerate(ab_bbox_result)
    ]
    labels_ab = np.concatenate(labels_ab)

    assert bboxes_rg.ndim == 2
    assert labels_rg.ndim == 1
    assert bboxes_rg.shape[0] == labels_rg.shape[0]
    assert bboxes_rg.shape[1] == 4 or bboxes_rg.shape[1] == 5

    assert bboxes_ab.ndim == 2
    assert labels_ab.ndim == 1
    assert bboxes_ab.shape[0] == labels_ab.shape[0]
    assert bboxes_ab.shape[1] == 4 or bboxes_ab.shape[1] == 5

    bboxes = np.vstack((rg_bbox_result + ab_bbox_result))
    labels = np.hstack((labels_rg, labels_ab))

    if score_thr > 0:
        assert bboxes.shape[1] == 5
        scores = bboxes[:, -1]
        inds = scores > score_thr
        bboxes = bboxes[inds, :]
        labels = labels[inds]
        scores = scores[inds]

    test_bboxes = py_cpu_nms(bboxes, labels, 0.5)
    new_bboxes = [bboxes[i] for i in test_bboxes]
    new_labels = [labels[i] for i in test_bboxes]
    new_scores = [scores[i] for i in test_bboxes]

    return new_bboxes, new_labels, new_scores


def load_name2mid(name2id):
    mmap = dict()
    source_wb = openpyxl.load_workbook(name2id, read_only=True)
    sheet = source_wb["Sheet1"]
    for row in sheet.rows:
        name = str(row[0].value)
        mid = str(row[1].value)
        eng = str(row[2].value)
        chn = str(row[3].value)
        mmap[name] = {
            "MID": mid,
            "ENG": eng,
            "CHN": chn
        }
    return mmap


def main():
    args = parse_args()

    rgcfg = mmcv.Config.fromfile(args.configrg)
    rgcfg.data.test.test_mode = True

    abcfg = mmcv.Config.fromfile(args.configab)
    abcfg.data.test.test_mode = True

    rg_config_file = args.configrg
    rg_checkpoint_file = args.checkpointrg

    ab_config_file = args.configab
    ab_checkpoint_file = args.checkpointab

    # build the model from a config file and a checkpoint file
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    modelrg = init_detector(rg_config_file, rg_checkpoint_file, device=device)
    modelrg.CLASSES = RoseGoldDataset.CLASSES

    modelab = init_detector(ab_config_file, ab_checkpoint_file, device=device)
    modelab.CLASSES = UltraABDataset.CLASSES

    acc = 0.0
    tot = 0.0
    gt_cls_num = np.zeros((len(modelrg.CLASSES) + len(modelab.CLASSES) - 1))
    tp = np.zeros((len(modelrg.CLASSES) + len(modelab.CLASSES) - 1))
    fp = np.zeros((len(modelrg.CLASSES) + len(modelab.CLASSES) - 1))
    fn = np.zeros((len(modelrg.CLASSES) + len(modelab.CLASSES) - 1))
    tn = np.zeros((len(modelrg.CLASSES) + len(modelab.CLASSES) - 1))

    cls2id = {
        '壳牌恒护超凡喜力欧系专属 5W-30 1L': 0,
        '壳牌恒护超凡喜力欧系专属 5W-30 4L': 1,
        '壳牌恒护超凡喜力欧系专属 5W-40 1L': 2,
        '壳牌恒护超凡喜力欧系专属 5W-40 4L': 3,
        '壳牌恒护超凡喜力亚系专属 5W-30 1L': 4,
        '壳牌恒护超凡喜力亚系专属 5W-30 4L': 5,
        '壳牌先锋超凡喜力 SN PLUS 天然气全合成油 0W-20 4L': 6,
        '壳牌先锋超凡喜力 SN PLUS 天然气全合成油 0W-20 1L': 7,
        '壳牌先锋超凡喜力 SN PLUS 天然气全合成油 0W-30 4L': 8,
        '壳牌先锋超凡喜力 SN PLUS 天然气全合成油 0W-30 1L': 9,
        '壳牌先锋超凡喜力 ACEA C5 天然气全合成油 0W-20 4L': 10,
        '壳牌先锋超凡喜力 ACEA C5 天然气全合成油 0W-20 1L': 11,
        '壳牌先锋超凡喜力 ACEA C2 / C3 天然气全合成油 0W-30 4L': 12,
        '壳牌先锋超凡喜力 ACEA C2 / C3 天然气全合成油 0W-30 1L': 13,
        '壳牌先锋超凡喜力 ACEA A3 / B4 天然气全合成油 0W-40 4L': 14,
        '壳牌先锋超凡喜力 ACEA A3 / B4 天然气全合成油 0W-40 1L': 15,
        '其他': 16
    }
    id2cls = {value: key for key, value in cls2id.items()}

    with open(test_path, "r") as f:
        filenames = f.readlines()
        for filename in filenames:
            img_file = filename.strip() + ".jpg"
            xml_file = filename.strip() + ".xml"
            img = cv2.imread(os.path.join(test_img_path, img_file))
            if img is not None:
                xml_path = os.path.join(test_xml_path, xml_file)
                coords = read_xml(xml_path)
                if len(coords) is 0:
                    print("No annotations\n")
                    continue
                gt_bboxes = [coord[:4] for coord in coords]
                gt_labels = [coord[4] for coord in coords]
                for label in gt_labels:
                    gt_cls_num[cls2id[label]] += 1
                    tot += 1
                resultrg = inference_detector(modelrg, img)
                resultab = inference_detector(modelab, img)
                det_bboxes, det_labels, det_scores = get_result(resultrg, resultab, modelrg.CLASSES, modelab.CLASSES, score_thr=0.5)
                ious = bbox_overlaps(np.array(det_bboxes), np.array(gt_bboxes))
                ious_max = ious.max(axis=1)
                ious_argmax = ious.argmax(axis=1)
                gt_matched_det = np.ones((len(gt_bboxes))) * -1
                det_matched_gt = np.ones((len(det_bboxes))) * -1
                gt_matched_scores = np.zeros((len(gt_bboxes)))
                for i in range(0, len(det_bboxes)):
                    if ious_max[i] > 0.5:
                        target_gt = ious_argmax[i]
                        if gt_matched_scores[target_gt] < det_scores[i]:
                            gt_matched_scores[target_gt] = det_scores[i]
                            gt_matched_det[target_gt] = i
                            det_matched_gt[i] = target_gt
                    else:
                        fp[cls2id[det_labels[i]]] += 1
                for i in range(0, len(det_matched_gt)):
                    gt = int(det_matched_gt[i])
                    if gt > -1:
                        if det_labels[i] == gt_labels[gt]:
                            tp[cls2id[det_labels[i]]] += 1
                            acc += 1
                        else:
                            fp[cls2id[det_labels[i]]] += 1
    for i in range(0, len(cls2id)):
        fn[i] = gt_cls_num[i] - tp[i]
        tn[i] = gt_cls_num.sum() - fn[i] - tp[i] - fp[i]

    print("accuracy: %f" % (acc / tot))
    mat = np.zeros((len(cls2id), len(TABLE_HEAD)))

    for i in range(0, len(cls2id)):
        mat[i][0] = i
        mat[i][1] = gt_cls_num[i]
        mat[i][2] = tp[i]
        mat[i][3] = fp[i]
        mat[i][4] = fn[i]
        mat[i][5] = tp[i] / (tp[i] + fp[i])
        mat[i][6] = tp[i] / (tp[i] + fn[i])
        print("%s: %.0f gt, %.0f det, %.0f tp, precision: %.6f, recall: %.6f" %
              (id2cls[i], gt_cls_num[i], tp[i] + fp[i], tp[i], tp[i] / (tp[i] + fp[i]), tp[i] / (tp[i] + fn[i])))\

    if os.path.exists("shell_statistics.xlsx"):
        os.remove("shell_statistics.xlsx")
    workbook = openpyxl.Workbook("shell_statistics.xlsx")
    sheet = workbook.create_sheet("sheet")
    sheet.append(TABLE_HEAD)
    for i in range(0, len(id2cls)):
        label = id2cls[i]
        sheet.append([label, "%.0f" % gt_cls_num[i], "%.0f" % tp[i], "%.0f" % fp[i], "%.0f" % fn[i],
                      "%.6f" % (tp[i] / (tp[i] + fp[i])), "%.6f" % (tp[i] / (tp[i] + fn[i]))])

    workbook.save("shell_statistics.xlsx")


if __name__ == "__main__":
    main()
