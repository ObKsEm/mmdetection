import argparse
import os

import cv2
import mmcv
import numpy as np
import pycocotools.mask as maskUtils
import torch

import openpyxl
from PIL import Image, ImageFont, ImageDraw
from mmcv import color_val, imwrite
from mmdet.core.evaluation.bbox_overlaps import bbox_overlaps

from mmdet.ops import nms

from mmdet.apis import init_detector, inference_detector, show_result
from mmdet.datasets.shell import ShellDataset
from mmdet.datasets.sku import SkuDataset
import xml.etree.ElementTree as ET

TABLE_HEAD = ["名称", "MID", "MID English description", "MID Chinese description", "样本个数", "识别个数", "准确率"]

test_img_path = "/home/lichengzhi/mmdetection/data/VOCdevkit/shell/2019.9.27/JPEGImages"
test_xml_path = "/home/lichengzhi/mmdetection/data/VOCdevkit/shell/2019.9.27/Annotations"
test_path = "/home/lichengzhi/mmdetection/data/VOCdevkit/shell/2019.9.27/ImageSets/Main/test.txt"


def parse_args():
    parser = argparse.ArgumentParser(description='MMDet test detector')
    parser.add_argument('config', help='test config file path')
    parser.add_argument('checkpoint', help='checkpoint file')
    parser.add_argument(
        '--eval',
        type=str,
        nargs='+',
        choices=['proposal', 'proposal_fast', 'bbox', 'segm', 'keypoints'],
        help='eval types')
    parser.add_argument('--show', action='store_true', help='show results')
    parser.add_argument('--tmpdir', help='tmp dir for writing some results')
    parser.add_argument(
        '--launcher',
        choices=['none', 'pytorch', 'slurm', 'mpi'],
        default='none',
        help='job launcher')
    parser.add_argument('--local_rank', type=int, default=0)
    parser.add_argument('--image', help='detect image file')
    args = parser.parse_args()
    return args


def read_xml(xml_path):
    tree = ET.parse(xml_path)
    root = tree.getroot()
    objs = root.findall('object')
    coords = list()
    for ix, obj in enumerate(objs):
        name = obj.find('name').text
        box = obj.find('bndbox')
        x_min = float(box[0].text)
        y_min = float(box[1].text)
        x_max = float(box[2].text)
        y_max = float(box[3].text)
        coords.append([x_min, y_min, x_max, y_max, name])
    return coords


def get_result(result, score_thr=0.5):
    bbox_result = result

    bboxes = np.vstack(bbox_result)
    labels = [
        np.full(bbox.shape[0], i, dtype=np.int32)
        for i, bbox in enumerate(bbox_result)
    ]
    labels = np.concatenate(labels)

    assert bboxes.ndim == 2
    assert labels.ndim == 1
    assert bboxes.shape[0] == labels.shape[0]
    assert bboxes.shape[1] == 4 or bboxes.shape[1] == 5
    if score_thr > 0:
        assert bboxes.shape[1] == 5
        scores = bboxes[:, -1]
        inds = scores > score_thr
        bboxes = bboxes[inds, :]
        labels = labels[inds]
        scores = scores[inds]

    test_bboxes = nms(bboxes, 0.5)
    new_bboxes = [bboxes[i] for i in test_bboxes[1]]
    new_labels = [labels[i] for i in test_bboxes[1]]
    new_scores = [scores[i] for i in test_bboxes[1]]

    return new_bboxes, new_labels, new_scores


def load_name2mid(name2id):
    mmap = dict()
    source_wb = openpyxl.load_workbook(name2id, read_only=True)
    sheet = source_wb["Sheet1"]
    for row in sheet.rows:
        name = str(row[0].value)
        mid = str(row[1].value)
        eng = str(row[2].value)
        chn = str(row[3].value)
        mmap[name] = {
            "MID": mid,
            "ENG": eng,
            "CHN": chn
        }
    return mmap


def main():
    args = parse_args()

    cfg = mmcv.Config.fromfile(args.config)
    cfg.data.test.test_mode = True

    config_file = args.config
    checkpoint_file = args.checkpoint

    # build the model from a config file and a checkpoint file
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model = init_detector(config_file, checkpoint_file, device=device)
    model.CLASSES = ShellDataset.CLASSES
    cls2id = dict(zip(model.CLASSES, range(0, len(model.CLASSES))))
    # test a single image and show the results
    acc = 0.0
    tot = 0.0
    gt_cls_num = np.zeros((len(model.CLASSES)))
    det_cls_num = np.zeros((len(model.CLASSES)))
    with open(test_path, "r") as f:
        delta_gt_cls_num = np.zeros((len(model.CLASSES)))
        delta_det_cls_num = np.zeros((len(model.CLASSES)))
        filenames = f.readlines()
        for filename in filenames:
            img_file = filename.strip() + ".jpg"
            xml_file = filename.strip() + ".xml"
            img = cv2.imread(os.path.join(test_img_path, img_file))
            if img is not None:
                xml_path = os.path.join(test_xml_path, xml_file)
                coords = read_xml(xml_path)
                if len(coords) is 0:
                    print("No annotations\n")
                    continue
                gt_bboxes = [coord[:4] for coord in coords]
                gt_labels = [coord[4] for coord in coords]
                for label in gt_labels:
                    gt_cls_num[cls2id[label]] += 1
                    delta_gt_cls_num[cls2id[label]] += 1
                tot += len(gt_bboxes)
                result = inference_detector(model, img)
                det_bboxes, det_labels, det_scores = get_result(result, score_thr=0.5)
                ious = bbox_overlaps(np.array(det_bboxes), np.array(gt_bboxes))
                ious_max = ious.max(axis=1)
                ious_argmax = ious.argmax(axis=1)
                gt_matched_det = np.ones((len(gt_bboxes))) * -1
                gt_matched_scores = np.zeros((len(gt_bboxes)))
                for i in range(0, len(det_bboxes)):
                    if ious_max[i] > 0.5:
                        target_gt = ious_argmax[i]
                        if gt_matched_scores[target_gt] < det_scores[i]:
                            gt_matched_scores[target_gt] = det_scores[i]
                            gt_matched_det[target_gt] = i
                for i in range(0, len(gt_matched_det)):
                    det = int(gt_matched_det[i])
                    if det is not -1:
                        if model.CLASSES[det_labels[det]] == gt_labels[i]:
                            det_cls_num[det_labels[det]] += 1
                            delta_det_cls_num[det_labels[det]] += 1
                            acc += 1
                assert (delta_det_cls_num <= delta_gt_cls_num).all()
    print("total gt: %d,  det: %d" % (tot, acc))
    print("accuracy: %f" % (acc / tot))
    mat = np.zeros((len(model.CLASSES), 4))
    for i in range(0, len(model.CLASSES)):
        print("%s: %.0f gt, %.0f det, accuracy: %.6f" %
              (model.CLASSES[i], gt_cls_num[i], det_cls_num[i], det_cls_num[i] / gt_cls_num[i] if gt_cls_num[i] > 0 else 0))
        mat[i][0] = i
        mat[i][1] = gt_cls_num[i]
        mat[i][2] = det_cls_num[i]
        mat[i][3] = det_cls_num[i] / gt_cls_num[i] if gt_cls_num[i] > 0 else 0
    print("----------------------------------------------------")
    mat = mat[(-mat[:, 1]).argsort()]
    inds = mat[:, 1] > 100
    pos = mat[inds]
    neg = mat[~inds]
    pos = pos[(-pos[:, 3]).argsort()]
    neg = neg[(-neg[:, 3]).argsort()]
    mmap = load_name2mid("Name2MID.xlsx")
    for i in range(0, len(pos)):
        item = pos[i]
        label = int(item[0])
        gt = int(item[1])
        det = int(item[2])
        acc = item[3]
        print("%s: %.0f gt, %.0f det, %.6f acc" % (model.CLASSES[label], gt, det, acc))
    print("----------------------------------------------------")
    for i in range(0, len(neg)):
        item = neg[i]
        label = int(item[0])
        gt = int(item[1])
        det = int(item[2])
        acc = item[3]
        print("%s: %.0f gt, %.0f det, %.6f acc" % (model.CLASSES[label], gt, det, acc))

    workbook = openpyxl.Workbook("statistics.xlsx")
    pos_sheet = workbook.create_sheet("positive")
    pos_sheet.append(TABLE_HEAD)
    for row in range(0, len(pos)):
        item = pos[row]
        label = model.CLASSES[int(item[0])]
        gt = int(item[1])
        det = int(item[2])
        acc = item[3]
        # pos_sheet.cell(row=row + 2, column=1, value=label)
        # pos_sheet.cell(row=row + 2, column=2, value=mmap[label]["MID"])
        # pos_sheet.cell(row=row + 2, column=3, value=mmap[label]["ENG"])
        # pos_sheet.cell(row=row + 2, column=4, value=mmap[label]["CHN"])
        # pos_sheet.cell(row=row + 2, column=5, value="%.0f" % gt)
        # pos_sheet.cell(row=row + 2, column=6, value="%.0f" % det)
        # pos_sheet.cell(row=row + 2, column=7, value="%.6f" % acc)
        pos_sheet.append([label, mmap[label]["MID"], mmap[label]["ENG"], mmap[label]["CHN"],
                          "%.0f" % gt, "%.0f" % det, "%.6f" % acc])

    neg_sheet = workbook.create_sheet("negative")
    for col in range(0, len(TABLE_HEAD)):
        # neg_sheet.cell(row=1, column=col + 1, value=TABLE_HEAD[col])
        neg_sheet.append(TABLE_HEAD)
    for row in range(0, len(neg)):
        item = neg[row]
        label = model.CLASSES[int(item[0])]
        gt = int(item[1])
        det = int(item[2])
        acc = item[3]
        # neg_sheet.cell(row=row + 2, column=1, value=label)
        # neg_sheet.cell(row=row + 2, column=2, value=mmap[label]["MID"])
        # neg_sheet.cell(row=row + 2, column=3, value=mmap[label]["ENG"])
        # neg_sheet.cell(row=row + 2, column=4, value=mmap[label]["CHN"])
        # neg_sheet.cell(row=row + 2, column=5, value="%.0f" % gt)
        # neg_sheet.cell(row=row + 2, column=6, value="%.0f" % det)
        # neg_sheet.cell(row=row + 2, column=7, value="%.6f" % acc)
        neg_sheet.append([label, mmap[label]["MID"], mmap[label]["ENG"], mmap[label]["CHN"],
                          "%.0f" % gt, "%.0f" % det, "%.6f" % acc])

    workbook.save("statistics.xlsx")


if __name__ == "__main__":
    main()
