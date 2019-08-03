import openpyxl
import numpy as np


B2C_list = "/Users/lichengzhi/bailian/壳牌/B2C SKU List.xlsx"
Name2MID = "/Users/lichengzhi/bailian/壳牌/Name2MID.xlsx"

MAP = {
    '壳牌喜力 HX6 SN 合成技术发动机油 5W-30 1L': 165,
    '壳牌喜力 HX6 SN 合成技术发动机油 5W-30 4L': 166,
    '壳牌喜力 HX6 合成技术发动机油 5W-40 1L': 162,
    '壳牌喜力 HX6 合成技术发动机油 5W-40 4L': 163,
    '壳牌喜力 HX7 PLUS 全合成发动机油 5W-20 1L': 132,
    '壳牌喜力 HX7 PLUS 全合成发动机油 5W-20 4L': 133,
    '壳牌喜力 HX7 PLUS 全合成发动机油 5W-30 1L': 129,
    '壳牌喜力 HX7 PLUS 全合成发动机油 5W-30 4L': 130,
    '壳牌喜力 HX7 PLUS 全合成发动机油 5W-40 1L': 126,
    '壳牌喜力 HX7 PLUS 全合成发动机油 5W-40 4L': 127,
    '壳牌喜力 HX8 SN 全合成发动机油 0W-20 1L': 95,
    '壳牌喜力 HX8 SN 全合成发动机油 0W-20 4L': 96,
    '壳牌喜力 HX8 全合成发动机油 0W-30 1L': 92,
    '壳牌喜力 HX8 全合成发动机油 0W-30 4L': 93,
    '壳牌喜力 HX8 全合成发动机油 0W-40 1L': 89,
    '壳牌喜力 HX8 全合成发动机油 0W-40 4L': 90,
    '壳牌喜力 HX5 PLUS 合成技术发动机油 10W-40 1L': None,
    '壳牌喜力 HX5 PLUS 合成技术发动机油 10W-40 4L': 180,
    '壳牌极净超凡喜力 SN 天然气全合成机油 0W-20 1L': 11,
    '壳牌极净超凡喜力 SN 天然气全合成机油 0W-20 4L': 12,
    '壳牌极净超凡喜力 ECT C2 / C3 天然气全合成机油 0W-30 1L': 27,
    '壳牌极净超凡喜力 ECT C2 / C3 天然气全合成机油 0W-30 4L': 28,
    '壳牌极净超凡喜力 天然气全合成机油 0W-40 1L': 2,
    '壳牌极净超凡喜力 天然气全合成机油 0W-40 4L': 3,
    '壳牌超凡喜力 天然气全合成机油 5W-30 1L': None,
    '壳牌超凡喜力 天然气全合成机油 5W-30 4L': None,
    '壳牌超凡喜力 天然气全合成机油 5W-40 1L': None,
    '壳牌超凡喜力 天然气全合成机油 5W-40 4L': None,
    '壳牌极净超凡喜力 RACING 天然气全合成机油 10W-60 1L': 19,
    '壳牌极净超凡喜力 RACING 天然气全合成机油 10W-60 4L': 20,
    '壳牌超凡喜力 ECT C3 全合成机油中超版 5W-30 1L': 283,
    '壳牌超凡喜力 ECT C3 全合成机油中超版 5W-30 4L': 284,
    '壳牌超凡喜力 全合成机油中超版 5W-40 1L': 278,
    '壳牌超凡喜力 全合成机油中超版 5W-40 4L': 279,
    '壳牌超凡喜力 SN 天然气全合成机油中超版 0W-20 1L': 289,
    '壳牌超凡喜力 SN 天然气全合成机油中超版 0W-20 4L': 290,
    '壳牌超凡喜力 ECT 天然气全合成机油中超版 0W-30 1L': 287,
    '壳牌超凡喜力 ECT 天然气全合成机油中超版 0W-30 4L': 288,
    '壳牌超凡喜力 PROFESSIONAL ABB 天然气全合成机油中超版 0W-40 1L': 285,
    '壳牌超凡喜力 PROFESSIONAL ABB 天然气全合成机油中超版 0W-40 4L': 286,
    '壳牌喜力城市启停 全合成发动机油 0W-30 1L': 333,
    '壳牌喜力动力巅峰 全合成发动机油 0W-40 1L': 332,
    '壳牌喜力混动先锋 全合成发动机油 0W-20 1L': 334,
    '壳牌极净超凡喜力 SN PLUS 天然气全合成油 0W-40 1L': 21,
    '壳牌极净超凡喜力 SN PLUS 天然气全合成油 0W-40 4L': 22,
    '壳牌极净超凡喜力 ECT C2 / C3 天然气全合成油 0W-30 1L': 27,
    '壳牌极净超凡喜力 ECT C2 / C3 天然气全合成油 0W-30 4L': 28,
    '壳牌极净超凡喜力 SN PLUS 天然气全合成油 X 0W-30 1L': 24,
    '壳牌极净超凡喜力 SN PLUS 天然气全合成油 X 0W-30 4L': 25,
    '壳牌极净超凡喜力 SN PLUS 天然气全合成油 0W-16 1L': 33,
    '壳牌极净超凡喜力 SN PLUS 天然气全合成油 0W-16 4L': 34,
    '壳牌极净超凡喜力 SN PLUS 天然气全合成油 0W-20 1L': 30,
    '壳牌极净超凡喜力 SN PLUS 天然气全合成油 0W-20 4L': 31,
    '壳牌超凡喜力 SN PLUS 天然气全合成油 5W-30 1L': 47,
    '壳牌超凡喜力 SN PLUS 天然气全合成油 5W-30 4L': 48,
    '壳牌超凡喜力 SN PLUS 天然气全合成油 5W-40 1L': 41,
    '壳牌超凡喜力 SN PLUS 天然气全合成油 5W-40 4L': 42,
    '壳牌超凡喜力 ETC C3 天然气全合成油 5W-30 1L': 50,
    '壳牌超凡喜力 ETC C3 天然气全合成油 5W-30 4L': 51,
    '壳牌超凡喜力 SN PLUS 天然气全合成油 X 0W-30 1L': 44,
    '壳牌超凡喜力 SN PLUS 天然气全合成油 X 0W-30 4L': 45,
    '壳牌喜力 HX5 PLUS SN 合成技术润滑油 5W-30 4L': 183,
    '壳牌喜力 HX5 PLUS SN 合成技术润滑油 10W-40 4L': 181,
    '壳牌喜力 HX6 SN 合成技术润滑油 5W-30 1L 紫': None,
    '壳牌喜力 HX6 SN 合成技术润滑油 5W-30 4L 紫': None,
    '壳牌喜力 HX6 SN 合成技术润滑油 5W-30 1L 黄': 171,
    '壳牌喜力 HX6 SN 合成技术润滑油 5W-30 4L 黄': 172,
    '壳牌喜力 HX6 SN 合成技术润滑油 5W-40 1L': 173,
    '壳牌喜力 HX6 SN 合成技术润滑油 5W-40 4L': 174,
    '壳牌喜力 HX7 PLUS SN PLUS 全合成润滑油 5W-20 1L': 159,
    '壳牌喜力 HX7 PLUS SN PLUS 全合成润滑油 5W-20 4L': 160,
    '壳牌喜力 HX7 PLUS SN PLUS 全合成润滑油 5W-30 1L': 156,
    '壳牌喜力 HX7 PLUS SN PLUS 全合成润滑油 5W-30 4L': 157,
    '壳牌喜力 HX7 PLUS SN PLUS 全合成润滑油 5W-40 1L': 153,
    '壳牌喜力 HX7 PLUS SN PLUS 全合成润滑油 5W-40 4L': 154,
    '壳牌喜力 HX8 SN PLUS 先进全合成油 0W-20 1L': 114,
    '壳牌喜力 HX8 SN PLUS 先进全合成油 0W-20 4L': 115,
    '壳牌喜力 HX8 SN PLUS 先进全合成油 0W-40 1L': 108,
    '壳牌喜力 HX8 SN PLUS 先进全合成油 0W-40 4L': 109,
    '壳牌喜力 HX8 SN PLUS 先进全合成油 5W-30 1L': 120,
    '壳牌喜力 HX8 SN PLUS 先进全合成油 5W-30 4L': 121,
    '壳牌喜力 HX8 SN PLUS 先进全合成油 5W-40 1L': 117,
    '壳牌喜力 HX8 SN PLUS 先进全合成油 5W-40 4L': 118,
    '壳牌喜力 HX8 SN PLUS 先进全合成油 X 0W-30 1L': 111,
    '壳牌喜力 HX8 SN PLUS 先进全合成油 X 0W-30 4L': 112,
    '壳牌爱德王子城市穿梭 全合成发动机油 10W-40 1L': None,
    '壳牌爱德王子动力巅峰 全合成发动机油 15W-50 1L': None,
    '壳牌爱德王子水平对置 全合成发动机油 5W-40 1L': None,
    '壳牌超凡喜力 全合成润滑油中超限量版 5W-30 1L': 281,
    '壳牌超凡喜力 全合成润滑油中超限量版 5W-30 4L': 282,
    '壳牌超凡喜力 全合成润滑油中超限量版 5W-40 1L': 276,
    '壳牌超凡喜力 全合成润滑油中超限量版 5W-40 4L': 277,
    '壳牌超凡喜力 全合成润滑油 5W-40 1L': 35,
    '壳牌超凡喜力 全合成润滑油 5W-40 4L': 36,
    '壳牌极净超凡喜力 天然气全合成润滑油 0W-20 1L': 11,
    '壳牌极净超凡喜力 天然气全合成润滑油 0W-20 4L': 12,
    '壳牌极净超凡喜力 天然气全合成润滑油 0W-30 1L': 8,
    '壳牌极净超凡喜力 天然气全合成润滑油 0W-30 4L': 9,
    '壳牌极净超凡喜力 天然气全合成润滑油 0W-40 1L': None,
    '壳牌极净超凡喜力 天然气全合成润滑油 0W-40 4L': None,
    '壳牌极净超凡喜力 天然气全合成润滑油 10W-60 4L': None,
    '壳牌劲霸 RIMULA 重负荷柴油机润滑油 R2 EXTRA': None,
    '壳牌劲霸 RIMULA 重负荷柴油机润滑油 R2': None,
    '壳牌劲霸 RIMULA R2 EXTRA 桶装': None,
    '壳牌劲霸 RIMULA 天然气发动机润滑油 R5 NG 10W-40': 381,
    '壳牌劲霸 RIMULA 天然气发动机润滑油 R5 NG 10W-40 桶装': 383,
    '壳牌劲霸 RIMULA 重负荷柴油机润滑油 R5 E 10W-40': 361,
    '壳牌劲霸 RIMULA 重负荷柴油机润滑油 R5 E 10W-40 桶装': 362,
    '壳牌劲霸 RIMULA 天然气发动机润滑油 R3 NG': None,
    '壳牌劲霸 RIMULA 重负荷柴油机润滑油 R3 TURBO': None,
    '壳牌劲霸 RIMULA 重负荷柴油机润滑油 R4 PLUS增强型 15W-40': 375,
    '壳牌劲霸 RIMULA 重负荷柴油机润滑油 R6 LM 10W-40': 403,
    '壳牌劲霸 RIMULA 重负荷柴油机润滑油 R6 M 10W-40': 386,
    '壳牌劲霸 RIMULA 重负荷柴油机润滑油 R4 X': None,
    '壳牌劲霸 RIMULA  R2 EXTRA 桶装': None,
    '壳牌劲霸 RIMULA  R3 TURBO 桶装': None,
    '壳牌爱德王子 AX3 摩托车润滑油 5W-30 1L': None,
    '壳牌爱德王子 AX3 摩托车润滑油 10W-30 1L': None,
    '壳牌爱德王子 AX3 摩托车润滑油 20W-50 1L': None,
    '壳牌爱德王子 AX3 摩托车润滑油 15W-40 1L': None,
    '壳牌爱德王子 AX3 重负荷摩托车发动机油 20W-50 1L': None,
    '壳牌爱德王子 ULTRA 全合成摩托车润滑油 5W-40 1L': None,
    '壳牌爱德王子 ULTRA 全合成摩托车润滑油 15W-50 1L': None,
    '壳牌爱德王子 AX5 合成技术摩托车润滑油 10W-40 1L': None,
    '壳牌爱德王子 AX7 半合成摩托车润滑油 10W-40 1L': None,
    '壳牌超凡喜力 ECT C3 天然气全合成机油 5W-30 1L': 38,
    '壳牌超凡喜力 ECT C3 天然气全合成机油 5W-30 4L': 39,
    '壳牌喜力 HX5 SN 优质多级润滑油 5W-30 1L': 188,
    '壳牌喜力 HX5 SN 优质多级润滑油 5W-30 4L': 189,
    '壳牌喜力 HX5 优质多级润滑油 10W-40 1L': 185,
    '壳牌喜力 HX5 优质多级润滑油 10W-40 4L': 186,
    '壳牌喜力 HX6 合成技术润滑油 10W-40 1L': 168,
    '壳牌喜力 HX6 合成技术润滑油 10W-40 4L': 169,
    '壳牌喜力 HX6 SN 合成技术润滑油 5W-30 1L': 176,
    '壳牌喜力 HX6 SN 合成技术润滑油 5W-30 4L': 177,
    '壳牌喜力 HX7 SN 合成技术润滑油 5W-30 4L': 138,
    '壳牌喜力 HX7 合成技术润滑油 5W-40 1L': 136,
    '壳牌喜力 HX7 合成技术润滑油 5W-40 4L': 135,
    '壳牌喜力 HX8 全合成润滑油 0W-20 1L': 95,
    '壳牌喜力 HX8 全合成润滑油 0W-20 4L': 96,
    '壳牌喜力 HX8 全合成润滑油 5W-30 4L': 102,
    '壳牌喜力 HX8 全合成润滑油 5W-40 1L': 98,
    '壳牌喜力 HX8 全合成润滑油 5W-40 4L': 99,
    '壳牌喜力 HX3 优质矿物润滑油 15W-40 4L': 194,
    '壳牌全效防冻液 OAT -30℃': None,
    '壳牌机动车发动机冷却液 OAT -30℃': None,
    '壳牌机动车发动机冷却液 OAT -45℃': None,
    '壳牌清洗油': 226,
    '未识别SKU': None
}


def main():
    sku_list = list()
    wb = openpyxl.load_workbook(B2C_list, read_only=True)
    sheet = wb.get_sheet_by_name("Sheet1")
    for row in sheet.rows:
        mid = str(row[7].value)
        eng = str(row[8].value)
        chn = str(row[9].value)
        sku_list.append({
            "MID": mid,
            "MID English description": eng,
            "MID Chinese description": chn
        })

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.cell(row=1, column=1, value="name")
    sheet.cell(row=1, column=2, value="MID")
    sheet.cell(row=1, column=3, value="MID English description")
    sheet.cell(row=1, column=4, value="MID Chinese description")
    row = 2
    k = np.array(list(MAP.keys()))
    v = np.array(list(MAP.values()))
    invalid = np.iinfo(np.int16).max
    for i in range(0, len(v)):
        if v[i] is None:
            v[i] = invalid
    m = np.dstack((k, v))[0]
    m = m[m[:, 1].argsort()]
    print(m)
    for i in range(0, m.shape[0]):
        name = m[i][0]
        line = m[i][1]
        sheet.cell(row=row, column=1, value=name)
        if line < invalid:
            sheet.cell(row=row, column=2, value=sku_list[line - 1].get("MID"))
            sheet.cell(row=row, column=3, value=sku_list[line - 1].get("MID English description"))
            sheet.cell(row=row, column=4, value=sku_list[line - 1].get("MID Chinese description"))
        else:
            sheet.cell(row=row, column=2, value="Unknown")
        row += 1
    workbook.save(Name2MID)


if __name__ == "__main__":
    main()

