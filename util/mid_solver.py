import openpyxl


source_dir = "/Users/lichengzhi/bailian/壳牌/Name2MID.xlsx"
target_dir = "/Users/lichengzhi/bailian/壳牌/SKU识别率统计说明.xlsx"


def main():
    mmap = dict()
    source_wb = openpyxl.load_workbook(source_dir, read_only=True)
    sheet = source_wb.get_sheet_by_name("Sheet1")
    for row in sheet.rows:
        name = str(row[0].value)
        mid = str(row[1].value)
        eng = str(row[2].value)
        chn = str(row[3].value)
        mmap[name] = {
            "MID": mid,
            "ENG": eng,
            "CHN": chn
        }
    target_wb = openpyxl.load_workbook(target_dir)
    sheet = target_wb["SKU准确率识别结果"]
    for row in range(1, len(tuple(sheet.rows))):
        name = str(sheet[row][5].value)
        if mmap.get(name, None) is not None:
            _ = sheet.cell(row=row, column=2, value=mmap[name]["ENG"])
            _ = sheet.cell(row=row, column=3, value=mmap[name]["CHN"])
            _ = sheet.cell(row=row, column=4, value=mmap[name]["MID"])

    target_wb.save(target_dir)


if __name__ == "__main__":
    main()
